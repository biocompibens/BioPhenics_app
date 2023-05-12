import warnings
from os import path

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Side, Alignment
from openpyxl.styles.numbers import FORMAT_NUMBER_00
from openpyxl.utils import cell
from pandas import (DataFrame, Series, ExcelWriter, merge,
                    concat, to_numeric)
from pandas.errors import MergeError

from scripts.combine_information import PreNormalization, sort_column
from scripts.fct_utils import (merge_rep, unique, compute_cluster,
                               is_deep_learning_feature)
from scripts.global_name import (DEEP_LEARNING_FEATURES, REPLICAT, LIGNEE,
                                 PLATE, WELLS, BARCODE, CONTENT, LABEL, SIRNA,
                                 FIELDS, PATH, WAVE, NOT_FEATURE, PLATE_AND_WELLS,
                                 CONTEXTE, COL_ORDERED, CLUSTER, SHEETNAMES)
from scripts.pandas_excel_styler import Styliste, DEFAULT_STYLE
from scripts.parameters import Parameters
from scripts.methods import Method


def combiner(other, this):
    if all(this.isnull()):
        return other
    else:
        return this


class Normalization(object):
    """Normalization object. It perform the actual normalization by calling
    :func:run method. """

    def __init__(self, pre_norm, parms=None):
        if isinstance(pre_norm, PreNormalization):
            self.pre_norm = pre_norm

        elif path.isfile(pre_norm):
            self.pre_norm = PreNormalization(unified_file=pre_norm)
        else:
            raise TypeError(
                f"Wrong instance for {pre_norm} : need to be an unified file")

        self.old_features = {}

        if parms is not None:
            if not isinstance(parms, Parameters):
                try:
                    parms = Parameters(parms)
                except BaseException:
                    raise TypeError(f"Error with your instance of parameters."
                                    f" Need a Parameters object (this errors can be raised "
                                    f"when Parameters is imported from another place "
                                    f"than in Screening_analysis_pipeline directory)")

            self.parms = parms
            if not self.parms.is_norm_ready(content=False):
                raise ValueError("Your parameters object is incomplete")
        else:
            self.parms = Parameters()

        if self.parms.other_parms.get('reduced_feature_space', False):
            self.pre_norm.reduce_feature_space(**self.parms.other_parms.get('reduced_feature_space_parms', {}))

        self.run_ready = self.parms.is_norm_ready(content=False)

        self.img_path_outlier = None

        self.hits_by_replicate = None
        self.merged_hitlist = None

        self.correction = Method('spatial_correction')
        self.normalisation = Method('normalization')
        self.hit_selection = Method('selection')

        self._transformed = None
        self._corrected = None
        self._normalized = None

    @property
    def data(self):
        return self.pre_norm.data

    @data.setter
    def data(self, df):
        self.pre_norm.data = df

    def median_feature(self, feats=None, selection_feature=True):
        if feats is None:
            self.replace_deep_learning_features(selection_feature=selection_feature)
            feats = self.parms.features
        if not isinstance(feats, list):
            feats = [feats]
        result = [f'median of {f}' for f in feats]
        self.change_back_features_names()
        return result

    @property
    def replicat_name(self):
        return self.data[REPLICAT].unique()

    @property
    def corrected(self):
        if self._corrected is None or self._corrected.empty:
            return self.data
        return self.data.combine(self._corrected, combiner)

    @corrected.setter
    def corrected(self, value):
        self._corrected = value

    @property
    def normalized(self):
        if self._normalized is None or self._normalized.empty:
            return self.corrected
        return self.data.combine(self._normalized, combiner)

    @normalized.setter
    def normalized(self, value):
        self._normalized = value

    def replace_deep_learning_features(self, parms=None, selection_feature=False):
        if parms is None:
            parms = self.parms
        for i, p in enumerate(iter(parms)):
            if self.is_considered_deep_learning_exp(p):
                self.old_features[i] = p.features[:]
                if selection_feature:
                    p.features = p.selection_features
                else:
                    p.features.remove(DEEP_LEARNING_FEATURES)
                    p.features += [col for col in self.pre_norm.data.columns if
                                   is_deep_learning_feature(col)]
        return parms

    def change_back_features_names(self, parms=None):
        if parms is None:
            parms = self.parms
        for i, p in enumerate(parms):
            if i in self.old_features:
                p.features = self.old_features[i]
        return parms

    def validation_outliers(self, data=None, t=-1):
        """
        This function compare hits in each replicate and return
        only those who are present in t replicate
        if nb_rep > t > 0 else all replicates

        Parameters
        ----------
        data: DataFrame or None
            data to validate
        t: int,
            minimal number of replicate in which a hit must be found

        Returns
        -------

        hits: DataFrame,
            validated hits, sorted by median of features

        """

        if data is None:
            data = self.hits_by_replicate

        if not 0 < t < self.pre_norm.replicate_number:
            t = self.pre_norm.replicate_number

        if data is None:
            raise ValueError(
                "You have to run a normalization "
                "with a hit selection method to have hits"
            )

        on_cols = [LIGNEE, PLATE, WELLS]
        # removed use of transform because in case of na value in on_cols,
        # transform doesn't work (if groupby dropna=False => throw an error
        # if groupby dropna=True => doesn't return all index of data)

        result = data.copy()

        groups = data.groupby(
            on_cols, sort=False, dropna=False
        )
        for idx, df in groups[LABEL]:
            lc = pd.concat([result[i[0]] == i[1] for i in zip(on_cols, idx)], axis=1).all(axis=1)
            result.loc[lc, LABEL] = (df.sum() >= t)

        return result[LABEL]

    def is_considered_deep_learning_exp(self, parms=None):
        if parms is None:
            parms = self.parms
        return self.pre_norm.deep_learning and DEEP_LEARNING_FEATURES in parms.features

    @staticmethod
    def sirna(hitlist):
        """ Add a column that count each hit by compound
        """

        def count_hit(x):
            nb = x[LABEL].astype(int).value_counts()
            try:
                nb_hit = nb.loc[1]
            except KeyError:
                nb_hit = 0
            nb_total = nb.sum()
            return Series(f"{nb_hit:.0f}/{nb_total:.0f}", index=x.index)

        result = hitlist.groupby([LIGNEE, REPLICAT, CONTENT],
                                 sort=False).apply(count_hit)
        result.index = result.index.droplevel([0, 1, 2])
        hitlist[SIRNA] = result

        return hitlist

    def run(self, parameters=None, reccord_parms=True):
        """
        This function will launch the analysis with parameters

        Parameters
        ----------

        parameters: Parameters
            instance of :class:`~internal_script.parameters.Parameters`
            with all necessary parameters for the analysis to be launched
        reccord_parms: bool
            if true parameter are reccord in this norm in parms atrribute

        Returns
        -------

        Outliers : A list of outliers found with this analysis

        """
        if parameters is None:
            parameters = self.parms
        else:
            if not isinstance(parameters, Parameters):
                parameters = Parameters(parameters)
        self.run_ready = parameters.is_norm_ready(content=False)

        if self.run_ready:

            corrected = DataFrame()
            normalized = DataFrame()
            hits_by_replicate = DataFrame()

            parameters = self.replace_deep_learning_features(parameters)

            for p in parameters:
                cor = self.correction(self.data, p)

                corrected = concat([corrected, cor], axis=1)
                self.corrected = cor

                nor = self.normalisation(self.corrected, p)
                normalized = concat([normalized, nor], axis=1)
                self.normalized = nor

                hi = self.hit_selection(self.normalized, p)

                if p.selection:
                    hi[LABEL] = self.validation_outliers(hi, t=p.validation)

                    if LABEL not in hits_by_replicate:
                        hits_by_replicate = hi.copy()
                    else:
                        hits_by_replicate[LABEL] = hits_by_replicate[
                                                       LABEL] & hi[LABEL]
                        hits_by_replicate[p.features] = hi[p.features]
                        added_col = [col for col in hi if col not in hits_by_replicate]
                        if added_col:
                            hits_by_replicate[added_col] = hi[added_col]
                    # make an intersection between analysis

            self.corrected = corrected
            self.normalized = normalized

            if not hits_by_replicate.empty:
                self.hits_by_replicate = hits_by_replicate

            if self.have_hit():
                if self.pre_norm.is_sirna():
                    self.hits_by_replicate = self.sirna(self.hits_by_replicate)
                self.merged_hitlist = self.format_hitlist()

            self.img_path_outlier = self.get_outliers_img_path()

            parameters = self.change_back_features_names(parameters)

            if reccord_parms:
                self.parms = parameters

            return self.hits_by_replicate

    def have_hit(self):
        """ Return the number of hit if at least
        one hits is found and False otherwise """
        if self.hits_by_replicate is not None:
            return self.hits_by_replicate[
                       LABEL].sum() / self.pre_norm.replicate_number
        return False

    def get_outliers_img_path(self):
        """
        Get a list of path corresponding to picture from wells

        Returns
        -------
        img_path_outlier : Pandas dataframe of a list of picture path
        """
        if self.pre_norm.img_path is not None and self.have_hit():
            img_path_outlier = merge(self.pre_norm.img_path,
                                     self.hits_by_replicate,
                                     on=[BARCODE, WELLS])[
                [REPLICAT, CONTENT, LIGNEE,
                 PLATE, BARCODE, WELLS, FIELDS,
                 WAVE, PATH]]

            img_path_outlier = img_path_outlier.set_index(
                [REPLICAT, LIGNEE, PLATE,
                 BARCODE, CONTENT, WELLS, FIELDS, WAVE])

        else:
            img_path_outlier = None

        return img_path_outlier

    def format_hitlist(self, hitlist=None):
        """
        Merge :const:BARCODE and features columns by replicate and compute
        the median of each feature
        """
        hitlist = hitlist if hitlist is not None else self.hits_by_replicate
        hit_index = [PLATE, WELLS, CONTENT]
        if SIRNA in hitlist.columns:
            hit_index.append(SIRNA)

        on_cols = [LIGNEE, LABEL] + hit_index

        try:
            hits = merge_rep(hitlist, REPLICAT, on_cols)
        except MergeError:
            raise MergeError('Error in isolating plates. You may have multiple plates '
                             'with same name, same concentration (if available) but different values')
        for f in hitlist.columns:
            if f not in NOT_FEATURE:
                inter = []
                replicated_names = [f + '_' + rn for rn in self.replicat_name]
                for c in hits.columns:
                    if c in replicated_names:
                        inter.append(c)
                hits[self.median_feature(f)[0]] = hits[inter].median(axis=1)
        return hits

    def _format_hitlist_to_excel(self, add_cluster_group=False,
                                 keep=False,
                                 renamed_colname=None, select_hit=True,
                                 filtered_sirna=True,
                                 color_threshold=0.7):
        """
        Format the hit page to create an excel sheet with the hitlist
        """
        to_drop = [LABEL]
        df = self.merged_hitlist if self.merged_hitlist is not None else self.format_hitlist()

        ctrl = df[CONTENT].isin(self.parms.ctrl)
        hits = df[df | ctrl] if select_hit else df

        hits = hits.drop(columns=to_drop)

        if add_cluster_group:
            hits = hits.merge(
                compute_cluster(hits, self.median_feature(),
                                color_threshold=color_threshold),
                left_on=[LIGNEE, CONTENT], right_index=True
            )

        if renamed_colname:
            hits = hits.rename({feat_name: feat_name.replace(key, value)
                                for key, value in renamed_colname.items()
                                for feat_name in hits.columns
                                if key in feat_name}, axis=1)

        is_sirna_exp = self.pre_norm.is_sirna()
        if is_sirna_exp:
            def concat_plate_wells(x):
                pname = x[PLATE].unique()[0]
                list_wells = ', '.join(x[WELLS].values).strip(", ")
                return Series(f"{pname} ({list_wells})", index=x.index)

            result = hits.groupby(
                [PLATE, LIGNEE, CONTENT], sort=False
            ).apply(concat_plate_wells)

            result.index = result.index.droplevel([0, 1, 2])
            hits[PLATE_AND_WELLS] = result

            feat_by_rep_col = [col for feature in self.parms.selection_features
                               for col in hits.columns if feature in col
                               and SIRNA not in col and 'median' not in col]

            hits = hits.drop(columns=[PLATE, WELLS] + feat_by_rep_col)

        ctrl_index = hits.index[hits[CONTENT].isin(self.parms.ctrl)]

        def sirna_case(data, selection_feature):
            real_hit = data
            for s_f in selection_feature:
                if s_f in data.name:
                    nb_hit = 0
                    for atomic_int in self.parms.get_interval_of(s_f):
                        isin = atomic_int.isin(data)
                        # check if every value of data is in atomic_int
                        if isin.any() and not isin.all():
                            if isin.sum() > nb_hit:
                                # interval with more hits than the previous
                                nb_hit = isin.sum()
                                real_hit = data.loc[isin]
                            elif isin.sum() == nb_hit and nb_hit >= len(data)/3:
                                # same number of hit in two intervals
                                # and no other interval can have more hit
                                return data.median(), 0
                        elif isin.all():
                            nb_hit = isin.sum()
                    if not nb_hit:  # no hit for this feature
                        return data.median(), nb_hit
                    # take the second most potent sirna
                    sorted_data = real_hit.sort_values(key=lambda y: y.abs())
                    try:
                        return sorted_data.iloc[-2], nb_hit
                    except IndexError:
                        return sorted_data.iloc[-1], nb_hit
            else:
                return data.median(), 0

        def agg_func(x):
            if len(x.unique()) > 1 or is_sirna_exp:
                # particular case when only 1 siRNA is outliers
                if any(c_name in x.name for c_name in (
                        BARCODE, PLATE, WELLS, PLATE_AND_WELLS, CONTEXTE, SIRNA
                )):
                    return ', '.join(x.unique())
                elif x.dtype in ['float64', 'int64']:
                    if is_sirna_exp and not x.index.isin(ctrl_index).any():
                        vals, nb_hit = sirna_case(x, self.parms.selection_features)
                        if x.name.startswith('median'):
                            return f"{vals};;{nb_hit}"
                        return vals
                    else:
                        return x.median()
                else:
                    if len(x.unique()) != 1:  # last chance...
                        raise ValueError(f"non implemented columns {x.name}")
            return x.values[0]

        # change that to be groupby([LIGNEE, CONTENT]).agg(...)
        # then pivot
        # agg must take into account self.parms.s_parms_mixed['str_parms']
        # if relative == '><' (or multiple rule on same feature :
        # how can i check that)
        # => values must be on the same interval

        # fire some warning about removing invalid columns
        # will throw an error with newer version of pandas
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            hits = hits.pivot_table(index=[CONTENT], columns=LIGNEE,
                                    aggfunc=agg_func)
        if is_sirna_exp:
            for c in hits.columns:
                for f in self.parms.selection_features:
                    if f in c[0] and c[0].startswith('median'):
                        result = hits[c].astype(str).str.split(';;', expand=True)
                        hits[c] = to_numeric(result[0].fillna(hits[c]),
                                             errors="ignore")
                        if 1 in result:
                            hits[(SIRNA + f" ({f})", c[1])] = (
                                result[1] + '/' +
                                hits[(SIRNA, c[1])].str.split('/', expand=True)[1]
                            )

            hits = hits.drop([col for col in hits.columns if col[1] == SIRNA],
                             axis=1)
            if filtered_sirna:
                # filter SiRNA based on hits number for each feature independently (at least 2 siRNAs as hits
                # in 1 feature)
                hits = hits.loc[DataFrame(
                    [hits[col].str.split('/', expand=True)[0].astype(float) > 1
                     for col in hits.columns
                     if SIRNA in col[0] and col[0] != SIRNA]
                ).any() | hits.index.isin(self.parms.ctrl)]
        hits = hits.swaplevel(axis=1).sort_index(axis=1, level=0)

        base_col = []
        for col in hits.columns.levels[1]:
            for colname in COL_ORDERED:
                if colname in col and col not in base_col:
                    if SIRNA not in col or col == SIRNA:
                        base_col.append(col)
        # got a case where same column were count twice

        # multi-index name of feature
        feat_col = [col for col in hits.columns.levels[1]
                    if col not in base_col]

        if not keep:
            foi = list(self.parms.selection_features)  # will be edited otherwise
            foi += [f for f in self.parms.features if f not in foi]
            feat_col = unique([
                col for f in foi
                for col in feat_col if f in col
            ])
            if CLUSTER in hits.columns.get_level_values(1):
                feat_col += [CLUSTER]

        hits = hits.reindex(base_col + feat_col, axis=1, level=1)
        sort_by = [c for c in hits.columns
                   if any([f in c[1] for f in self.parms.selection_features])]
        if not sort_by:  # in case of ic50 exp with dss_lum for feature and
            # dss_percentage control_transformed_corrected_etc_lum
            # the first one is not working
            try:
                sort_by = [c for c in hits.columns
                           if any([f.rsplit('_', 1)[1] in c[1]
                                   for f in self.parms.selection_features])]
            except IndexError:
                pass
        if sort_by:
            hits = hits.sort_values(by=sort_by)
        nb_hit0 = len(hits.loc[~hits.index.isin(self.parms.ctrl)])

        # place ctrl in first
        hit_number = DataFrame(index=[f"Hit number: {nb_hit0}"], columns=hits.columns)
        ctrl_neg_name = DataFrame(columns=hits.columns, index=["Negative control"])
        ctrl_pos_name = DataFrame(columns=hits.columns, index=["Positive control"])
        empty = DataFrame(columns=hits.columns, index=[" "])
        hit_list = DataFrame(columns=hits.columns, index=["Hitlist"])

        result = concat([hit_number, ctrl_neg_name, hits.loc[hits.index.isin(self.parms.ctrl_neg)], empty])

        if self.parms.ctrl_pos:
            result = concat([result, ctrl_pos_name, hits.loc[hits.index.isin(self.parms.ctrl_pos)], empty])

        result = concat([result, hit_list, hits.loc[~hits.index.isin(self.parms.ctrl)]])
        result.index.name = CONTENT
        return result

    def to_excel(self, out, **kwargs):
        """
        Write an excel file

        Parameters
        ----------

        out : FileLike object, ExcelWriter or String (path)
            The file (or name) of the resulting output

        Returns
        -------
        ExcelWriter
        """

        features_interest = self.parms.features
        if not isinstance(out, ExcelWriter) or out.engine != 'openpyxl':
            out = ExcelWriter(out, engine="openpyxl")

        sheet = kwargs.get('sheetnames', SHEETNAMES)
        # list of 9 values

        include = kwargs.get('include', ['raw', 'norm'])
        # value can be 'raw', 'corr' and 'norm', 'median'.
        # Other are ignored

        filtered = kwargs.get('filtered', True)

        def prep_df(dataframe):
            result = sort_column(dataframe)
            return result[[
                c for c in result.columns if c in COL_ORDERED + features_interest
            ]]

        try:
            with out as writer:
                if 'raw' in include:
                    pre_norm = Styliste(sort_column(self.pre_norm.get_original_content()),
                                        writer, sheet_name=sheet[0], index=False,
                                        header_style={"alignment": Alignment(wrapText=True)})
                    pre_norm.write()

                renamed_colname = {col: col for col in self.data.columns
                                   if col in features_interest}

                if self.parms.spatial_correction:
                    renamed_colname.update(
                        {c: f'corrected_{v}' for c, v in renamed_colname.items()}
                    )

                if 'corr' in include:
                    corrected = prep_df(
                        self.pre_norm.get_original_content(self.corrected)
                    ).rename(renamed_colname, axis=1)
                    corrected.to_excel(writer, sheet[6], index=False)

                if 'img' in include and self.img_path_outlier is not None \
                        and not self.img_path_outlier.empty:
                    self.img_path_outlier.to_excel(writer, sheet[3], merge_cells=False)
                    # merged cell caused bugs with openpyxl writer

                if self.parms.normalization:
                    renamed_colname.update(
                        {c: f'{self.parms[c].normalization}_{v}' if self.parms[c].normalization else v
                         for c, v in renamed_colname.items()}
                    )

                if 'norm' in include:
                    normalized = prep_df(
                        self.pre_norm.get_original_content(self.normalized)
                    ).rename(renamed_colname, axis=1)

                    normalized = concat([
                        normalized,
                        self.hit_selection.additional_features
                    ], axis=1)

                    on_col = [col for col in COL_ORDERED if
                              col in normalized.columns and col not in (
                                  BARCODE, LIGNEE)]

                    normalized = merge_rep(normalized, LIGNEE, on_col)
                    normalized.to_excel(writer, sheet[1], index=False)

                clustering = kwargs.get('add_cluster_group', False)
                color_threshold = kwargs.get('color_threshold', None)
                # maybe find more condition when clustering is not
                # wishable
                keep = kwargs.get('keep_other_column', False)

                if 'median' in include:
                    median_df = self._format_hitlist_to_excel(
                        select_hit=False, keep=keep,
                        renamed_colname=renamed_colname
                    )
                    median_df.to_excel(writer, sheet[8])

                if self.have_hit():
                    if 'pooled' in include:
                        self.merged_hitlist.loc[
                            self.merged_hitlist[LABEL],
                            [col for col in self.merged_hitlist.columns if col in NOT_FEATURE] +
                            [col for f in self.parms.selection_features
                             for col in self.merged_hitlist.columns if f in col]
                        ].drop(LABEL, axis=1).sort_values(by=CONTENT).to_excel(writer, sheet[9], index=False)

                    hits = self._format_hitlist_to_excel(
                        clustering, keep,
                        color_threshold=color_threshold,
                        renamed_colname=renamed_colname,
                        filtered_sirna=filtered
                    )

                    idx_col = {
                        f: list(hits.columns).index(f) + 1
                        for f in hits.columns
                        if ('median' in f[1] and any(
                            [feat in f[1] for analysis in self.parms
                             for feat in analysis.selection_features]
                        )) or  # classical
                        f[1].split('_', 1)[-1] in features_interest  # ic50
                    }

                    sheet_hits = Styliste(hits, writer, sheet_name=sheet[2],
                                          merge_cells=False,
                                          all_table_style=dict(
                                              border=dict(
                                                  left=Side(border_style='thin',
                                                            color='ffffff'),
                                                  right=Side(border_style='thin',
                                                             color='ffffff'),
                                                  top=Side(border_style='thin',
                                                           color='ffffff'),
                                                  bottom=Side(border_style='thin',
                                                              color='ffffff')),
                                              number_format=FORMAT_NUMBER_00
                                          ),
                                          **DEFAULT_STYLE)
                    sheet_hits.write(best_fit=True)

                    for feat, i in idx_col.items():
                        col = cell.get_column_letter(i+1)
                        writer.sheets[sheet[2]].conditional_formatting.add(
                            f"{col}2:{col}{len(hits) + 2}",
                            ColorScaleRule(start_type='percentile', start_value=10,
                                           start_color='963634',
                                           mid_type='num', mid_value=0,
                                           mid_color='ffffff',
                                           end_type='percentile', end_value=90,
                                           end_color='7EC234')
                        )
                    if 'specific_line' in kwargs and kwargs['specific_line']:
                        foi = {
                            c for c in hits.columns.get_level_values(1)
                            if any([f in c for f in self.parms.selection_features]) and 'median' in c
                        }
                        df = hits.loc[:, (slice(None), list(foi))]

                        def replace_hit_name(x):
                            a = pd.Series(index=x.index, dtype=object)
                            a.loc[~pd.isna(x)] = x.loc[~pd.isna(x)].index
                            return a

                        res = df.apply(replace_hit_name).groupby(LIGNEE, axis=1).apply(
                            lambda x: x.apply(
                                lambda row: None if row.isna().all() else row.dropna().unique()[0], axis=1
                            )
                        )
                        res = res.sort_values(by=[c for c in res.columns], key=lambda col: col.isna())
                        res = res.loc[res.isna().any(axis=1)]

                        specific_hits = Styliste(res, writer, sheet_name=sheet[10], merge_cells=False, index=False)
                        specific_hits.write(best_fit=True)

        except IndexError:
            raise IndexError(f"Wrong format for sheetnames {sheet}")
        return out


if __name__ == '__main__':
    import sys
    with open(sys.argv[2], "r") as parms_file:
        pms = Parameters(parms_file.read())
    norm = Normalization(sys.argv[1], parms=pms)
    norm.run()
    norm.to_excel(sys.argv[3], include=['raw', 'norm', 'median'])
