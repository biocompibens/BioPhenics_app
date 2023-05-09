from itertools import chain

import pandas as pd
import numpy as np
from openpyxl.cell import Cell
from openpyxl.formatting import Rule

from openpyxl.utils import cell as op_cell
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment, Protection
from openpyxl.styles.numbers import FORMAT_NUMBER
from openpyxl.worksheet.table import TableStyleInfo, Table


class StyleCell(object):
    font: Font
    fill: PatternFill
    border: Border
    alignment: Alignment
    number_format: str = FORMAT_NUMBER
    protection: Protection

    authorized_args = {'font': Font, 'fill': PatternFill, 'border': Border, 'alignment': Alignment,
                       'number_format': str, 'protection': Protection}

    def __getitem__(self, item):
        if item in self.authorized_args:
            try:
                return getattr(self, item)
            except AttributeError:
                return self.authorized_args[item]()
        raise AttributeError(f"Unknown arg : {item}")

    def keys(self):
        return self.authorized_args.keys()

    def __init__(self, dict_=None, **kwargs):
        if isinstance(dict_, StyleCell):
            kwargs.update(dict_.__dict__)
        if isinstance(dict_, dict):
            kwargs.update(dict_)

        for arg, instance in self.authorized_args.items():
            if arg in kwargs:
                if not isinstance((val_arg := kwargs.get(arg, {})), instance):
                    if val_arg is not None:
                        val_arg = instance(**val_arg)
                setattr(self, arg, val_arg)

        if self.number_format == '{}':
            # correct default value
            self.number_format = FORMAT_NUMBER
        
    def update(self, dict_=None, **kwargs):
        self.__init__(dict_=dict_, **kwargs)

    def apply_on(self, cells):
        if isinstance(cells, Cell):
            cells = [cells]
        list_of_cells = chain.from_iterable(cells) if not isinstance(cells[0], Cell) else cells
        # when selecting a range in openpyxl a tuple of tuple of cells is returned

        for cell in list_of_cells:
            for arg in self.authorized_args:
                if hasattr(self, arg):
                    setattr(cell, arg, getattr(self, arg))

    def copy(self):
        return StyleCell(**self.__dict__)


DEFAULT_STYLE = dict(
    header_style=StyleCell(font=Font(size=12, bold=True, vertAlign=None, color='FF0000'),
                           border=dict(left=Side(border_style='thin', color='FFffff'),
                                       right=Side(border_style='thin', color='FFffff'),
                                       top=Side(border_style='thin', color='FFffff'),
                                       bottom=Side(border_style='thin', color='FFffff'))),
    header_index_style=StyleCell(alignment=Alignment(horizontal="left")),
    header_col_style=StyleCell(alignment=Alignment(wrap_text=True, text_rotation=90,
                                                   horizontal='center', vertical="center"),
                               font=Font(color='ffffff'), fill=PatternFill(bgColor="5a5a5a", fill_type='mediumGray'))
)


class ExcelCell(object):
    def __init__(self, df, startrow, startcol, merge_cells=True, index=True, headers=True):
        self.columns = df.columns
        self.index = df.index
        self.index_lvl = df.index.nlevels if index else 0
        self.columns_lvl = (df.columns.nlevels if merge_cells else 1) if headers else 0
        self.startrow = int(startrow)
        self.startcol = int(startcol)
        self.lastrow = self.startrow + len(self.index) + self.columns_lvl - 1
        self.lastcol = self.startcol + len(self.columns) + self.index_lvl - 1

    @property
    def first_cell(self):
        return self.pandas2excel(self.startrow, self.startcol)

    @property
    def last_cell(self):
        return self.pandas2excel(self.lastrow, self.lastcol)

    def get_all_cells(self):
        return f"{self.first_cell}:{self.last_cell}"

    @staticmethod
    def _get_axis_range(items, base_start, base_end):
        if items is None:
            start = base_start - 1
            end = base_end
        elif not items:
            raise ValueError()
        else:
            start = base_start + items[0]
            end = base_start + items[-1]
        return start, end

    def get_range(self, row=None, col=None):
        try:
            start_index, end_index = self._get_axis_range(
                row, self.startrow + self.columns_lvl, self.lastrow
            )
            start_letter, end_letter = self._get_axis_range(
                col, self.startcol + self.index_lvl, self.lastcol
            )
        except ValueError:
            return
        return f'{self.pandas2excel(start_index, start_letter)}:{self.pandas2excel(end_index, end_letter)}'

    def _get_axis(self, axis):
        return self.columns if axis else self.index

    def _get_range_index(self, slicing, axis=0):
        which_axis = self._get_axis(axis)
        try:
            return which_axis[slicing]
        except TypeError:
            idx_start = which_axis.get_loc(slicing.start) if slicing.start is not None else None
            idx_stop = which_axis.get_loc(slicing.stop) if slicing.stop is not None else None
            return which_axis[idx_start:idx_stop:slicing.step]

    def __getitem__(self, item):
        if not isinstance(item, tuple):
            item = item, slice(None, None)

        row_item, col_item = [it if isinstance(it, (list, slice, pd.Series, pd.DataFrame)) else [it] for it in item]

        if isinstance(row_item, (pd.Series, pd.DataFrame)):
            try:
                row_item = self.index[row_item.squeeze()].tolist()
            except AttributeError:
                raise ValueError(f'Multidimensionnal data for row in {item}')

        if isinstance(col_item, (pd.Series, pd.DataFrame)):
            try:
                col_item = self.columns[col_item.squeeze()].tolist()
            except AttributeError:
                raise ValueError(f'Multidimensionnal data for col in {item}')

        if self.is_consecutive(row_item, axis=0) and self.is_consecutive(col_item, axis=1):
            if isinstance(row_item, slice):
                row_item = [row_item.start if row_item.start is not None else self.index[0],
                            row_item.stop if row_item.stop is not None else self.index[-1]]
            if isinstance(col_item, slice):
                col_item = [col_item.start if col_item.start is not None else self.columns[0],
                            col_item.stop if col_item.stop is not None else self.columns[-1]]
            row_item = [self.index.get_loc(row) for row in row_item]
            col_item = [self.columns.get_loc(col) for col in col_item]
            if not row_item or not col_item:
                return []
            return [self.get_range(row=row_item, col=col_item)]

        if isinstance(row_item, slice):
            row_item = self._get_range_index(row_item, axis=0)
        if isinstance(col_item, slice):
            col_item = self._get_range_index(col_item, axis=1)

        return [f"{self.get_col(col)}{self.get_row(row)}"
                for col in col_item for row in row_item]

    def get_row(self, row):
        return self.startrow + self.columns_lvl + self.index.get_loc(row) + 1

    def get_col(self, col):
        return op_cell.get_column_letter(self.startcol + self.index_lvl + self.columns.get_loc(col) + 1)

    @staticmethod
    def pandas2excel(row, col):
        """excel indexing start at 1 (pandas 0)"""
        return f"{op_cell.get_column_letter(col + 1)}{row + 1}"

    def get_header_cells(self):
        end_row = self.startrow + self.columns_lvl - 1
        end_col = self.startcol + len(self.columns) + self.index_lvl - 1
        return f'{self.first_cell}:{self.pandas2excel(end_row, end_col)}'

    def get_index_cells(self):
        start_row = self.startrow + self.columns_lvl - 1
        start_col = self.startcol
        end_row = self.startrow + len(self.index) + self.columns_lvl - 1
        end_col = self.startcol + self.index_lvl - 1
        return f'{self.pandas2excel(start_row, start_col)}:{self.pandas2excel(end_row, end_col)}'

    def get_index_col(self):
        return [op_cell.get_column_letter(col + 1) for col in range(self.startcol, self.startcol + self.index_lvl)]

    def get_header_row(self):
        return [row + 1 for row in range(self.startrow, self.startrow + self.columns_lvl)]

    def is_consecutive(self, items, axis=0):
        which_axis = self._get_axis(axis)
        if isinstance(items, slice):
            if items.step is not None:
                return False
            try:
                if items.start is not None:
                    which_axis.get_loc(items.start)
                if items.stop is not None:
                    which_axis.get_loc(items.stop)
            except KeyError as e:
                raise KeyError(f"{e} not found in {'columns' if axis else 'index'}")
            return True

        if len(items) == 1 and items[0] in which_axis:
            return True

        try:
            iloc = sorted([which_axis.get_loc(item) for item in items])
        except KeyError as e:
            raise KeyError(f"{e} not found in {'columns' if axis else 'index'}")
        return sum(np.diff(iloc) == 1) >= len(items)

    def correct_blank_line(self, sheet):
        if self.columns.nlevels > 1:
            # issue #27772 pandas
            blank_lines = sheet[self.startrow + self.columns_lvl + 1]
            if (v := blank_lines[0].value) is not None:
                previous_cell = sheet[f"{op_cell.get_column_letter(self.startcol + 1)}"
                                      f"{self.startrow + self.columns_lvl}"]
                if not previous_cell.value:
                    previous_cell.value = v
                    blank_lines[0].value = None
            if all([c.value is None for c in blank_lines]):
                sheet.delete_rows(self.startrow + self.columns_lvl + 1)


class Styliste(object):
    A_FACTOR = 1.1
    P_FACTOR = 1.2

    def __init__(self, df, excel_writer, header_style=None, header_col_style=None, header_index_style=None,
                 conditionnal_style=None, cell_style=None, row_height=None,
                 col_width=None, all_table_style=None, conditionnal_rules=None, **kwargs):
        """"""
        self.header = kwargs.pop('header', True)
        self.index = kwargs.pop('index', True)
        self.na_rep = kwargs.pop('na_rep', '')
        self.sheet_name = kwargs.pop("sheet_name", "Sheet1")
        self.tables = None

        if not isinstance(excel_writer, pd.ExcelWriter):
            self.excel_writer = pd.ExcelWriter(excel_writer, engine='openpyxl')
        else:
            self.excel_writer = excel_writer

        self.df = df

        self.header_col_style = StyleCell({**(header_style or {}), **(header_col_style or {})})
        self.header_index_style = StyleCell({**(header_style or {}), **(header_index_style or {})})
        self.conditionnal_style = [(condition, StyleCell(style)) for condition, style in (conditionnal_style or [])]
        self.cell_style = {k: StyleCell(v) for k, v in cell_style.items()} if cell_style is not None else {}
        self.conditionnal_rules = [Rule(**kw) for kw in conditionnal_rules or []]

        if row_height is not None and not isinstance(row_height, dict):
            raise TypeError("'row_height' must be a dictionary")
        self.row_height = row_height or {}

        if col_width is not None and not isinstance(col_width, dict):
            raise TypeError("'col_width' must be a dictionary")
        self.col_width = col_width or {}

        self.kwargs = kwargs

        self.loc_cell = ExcelCell(df, kwargs.get('startrow', 0), kwargs.get('startcol', 0),
                                  merge_cells=self.kwargs.get('merge_cells', True),
                                  index=self.index, headers=self.header)

        if all_table_style:
            self.cell_style[self.loc_cell.get_all_cells()] = StyleCell(all_table_style)

    def save(self):
        self.excel_writer.save()

    def apply_header_style(self, sheet, header=True, index=True):
        if header:
            self.header_col_style.apply_on(sheet[self.loc_cell.get_header_cells()])
            for row in self.loc_cell.get_header_row():
                sheet.row_dimensions[row].height = 100  # self.header_col_style.height
        if index:
            self.header_index_style.apply_on(sheet[self.loc_cell.get_index_cells()])

    def set_column_width(self, sheet, columns, width):
        """Set the width of the given columns
        :param sheet: workbook sheet
        :param set|list|tuple columns: a single or a list/tuple/set of column name,
        index or letter to change their width
        :param int|float width: numeric positive value of the new width
        :return: self
        :rtype: StyleFrame
        """

        if not isinstance(columns, (list, pd.Index)):
            columns = [columns]
        try:
            width = float(width)
        except ValueError:
            raise TypeError('columns width must be numeric value')

        if width <= 0:
            raise ValueError('columns width must be positive')

        for column in columns:
            sheet.column_dimensions[self.loc_cell.get_col(column)].width = width

    def set_column_width_dict(self, sheet, col_width=None):
        """"""
        col_width = col_width or self.col_width
        if col_width is not None:
            for cols, width in col_width.items():
                self.set_column_width(sheet, cols, width)

    def best_fit(self, sheet, subset=None, index=False):
        if not subset:
            subset = self.df.columns

        if not all([s in self.df.columns for s in subset]):
            raise ValueError(f'columns in {subset} not found in data')
        self.col_width.update({
            col: (self.df[col].astype(str).str.len().quantile(0.95) + self.A_FACTOR) * self.P_FACTOR for col in subset
        })
        if index:
            df_index = self.df.index.to_frame().T if self.df.index.nlevels > 1 else {'': self.df.index}
            for c, index_col in zip(self.loc_cell.get_index_col(), df_index):
                sheet.column_dimensions[c].width = max(df_index[index_col].str.len())

    @staticmethod
    def set_row_height(sheet, rows, height):
        """ Set the height of the given rows"""

        if not isinstance(rows, (set, list, tuple, pd.Index)):
            rows = [rows]
        try:
            height = float(height)
        except ValueError:
            raise TypeError('rows height must be numeric value')

        if height <= 0:
            raise ValueError('rows height must be positive')
        for row in rows:
            try:
                row = int(row)
            except TypeError:
                raise TypeError("row must be an index")

            sheet.row_dimensions[row].height = height

    def set_row_height_dict(self, sheet):
        """"""
        if self.row_height is not None:
            for rows, height in self.row_height.items():
                self.set_row_height(sheet, rows, height)

    def apply_conditionnal_style(self, sheet, conditionnal_style=None):
        """
        apply a style based on a condition (only work on a row or columns at a time...)
        Parameters
        ----------
        sheet
        conditionnal_style

        Returns
        -------

        """
        conditionnal_style = conditionnal_style or self.conditionnal_style
        if conditionnal_style is not None:
            for condition, style in conditionnal_style:
                for loc_cell in self.loc_cell[condition]:
                    style.apply_on(sheet[loc_cell])

    def apply_style_by_index(self, sheet, cell_style=None):
        cell_style = cell_style or self.cell_style
        if cell_style:
            for loc_cell, style in cell_style.items():
                style.apply_on(sheet[loc_cell])

    def make_excel_table(self):
        medium_style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                                      showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        self.tables = Table(ref=self.loc_cell.get_all_cells(), displayName='table_1', tableStyleInfo=medium_style)

    def apply_conditionnal_rule(self, sheet, conditionnal_rules=None):
        conditionnal_rules = conditionnal_rules or self.conditionnal_rules
        if conditionnal_rules:
            for rule in conditionnal_rules:
                sheet.conditional_formatting.add(self.loc_cell.get_all_cells(), rule)

    def write(self, best_fit=False):
        self.df.to_excel(self.excel_writer, sheet_name=self.sheet_name, engine='openpyxl', header=self.header,
                         index=self.index, na_rep=self.na_rep, **self.kwargs)
        sheet = self.excel_writer.sheets[self.sheet_name]

        self.loc_cell.correct_blank_line(sheet)

        self.apply_header_style(sheet, self.header, self.index)
        self.apply_conditionnal_style(sheet)
        self.apply_style_by_index(sheet)
        if best_fit:
            self.best_fit(sheet, index=self.index)
        self.set_column_width_dict(sheet)
        self.set_row_height_dict(sheet)
        if self.tables:
            sheet.add_table(self.tables)
        self.apply_conditionnal_rule(sheet)

    def close(self):
        return self.excel_writer.close()
